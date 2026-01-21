#  Importação de pacotes
import json
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Função responsável por configurar os arquivos de log (info.log e automation.log)
def configurar_logs(info_log, automation_log):
    # controle dos logs
    log = logging.getLogger()
    # Avisos dos logs -- Registros dos logs
    log.setLevel(logging.DEBUG)
    # evita logs duplicados
    log.handlers.clear()

    # Formatação dos logs Data/Horario
    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%d/%m/%Y %H:%M:%S"
    )

    # Informações  dos log
    info_handler = logging.FileHandler(info_log)
    info_handler.setLevel(logging.INFO)
    info_handler.setFormatter(formatter)

    # AUTOMATION log (tudo)
    automation_handler = logging.FileHandler(automation_log)
    automation_handler.setLevel(logging.DEBUG)
    automation_handler.setFormatter(formatter)

    log.addHandler(info_handler)
    log.addHandler(automation_handler)

    logging.info("Sistema de logs inicializado")

# Funçao de geração do arquivo JSON -- 
def salvar_json(dados, arquivo):
    try:
        with open(arquivo, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=4)
        logging.info("Arquivo JSON gerado com sucesso")
    except Exception as e:
        logging.critical(f"Falha ao gerar JSON: {e}")
        raise

# Função de geração da Planilha
def gerar_excel_do_json(arquivo_json, arquivo_excel):
    try:
        with open(arquivo_json, "r", encoding="utf-8") as f:
            dados = json.load(f)

        wb = Workbook() #Workbook cria o arquivo excel
        ws = wb.active
        ws.title = "Resultados Google Maps"

        cabecalhos = ["Nome", "Tipo", "Nota", "Quantidade Avaliações", "Endereço"]
        ws.append(cabecalhos)

        for item in dados:
            ws.append([
                item.get("nome"),
                item.get("tipo"),
                item.get("nota"),
                item.get("avaliacoes"),
                item.get("endereco")
            ])

        # Define e ajusta colunas do Excel
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

        wb.save(arquivo_excel)
        logging.info("Planilha Excel gerada com sucesso")

    except Exception as e:
        logging.critical(f"Erro ao gerar Excel: {e}")
        raise
